import pandas as pd
import time
import os
import requests
from secrets import Credentials
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font



# region getNewsAsDataFrame
def getNewsAsDataFrame(category='business'):
    url = f'https://newsapi.org/v2/top-headlines?country=us&category={category}&apiKey={Credentials.newsapiKey}'
    newsJson = requests.get(url).json()
    articles = newsJson['articles']
    titles = []
    urls = []
    for i, article in enumerate(articles):
        titles.append(article['title'])
        urls.append(article['url'])
    return pd.DataFrame({'TITLE': titles, 'URL': urls})
# endregion

# region write to xlsx
def writeNewsToFile(titles=[], urls=[]):
    # region csv code
    # with open(f'./CSVs/{category}_category_news_{time.strftime("%Y%m%d-%H%M%S")}.xlsx', 'a', encoding='UTF8', newline='') as f:
    #     w = csv.writer(f)
    #     w.writerow(['TITLE','.', 'URL'])
    #     for title,url in newsLst:
    #         w.writerow((title,'.',url))
    # endregion
    # convert to pandas df
    pass
# endregion


def addHyperlinks(filename):
    wb = load_workbook(filename)
    # go thru all sheets and apply hyperlink formula to every row in column B
    for i, sheetname in enumerate(wb.sheetnames):
        ws = wb[sheetname]
        # 1 to n because we're skipping the header
        n = len(wb[sheetname][get_column_letter(1)])
        # make a header for the new column
        ws.cell(row=1, column=3).value = 'Hyperlink'
        # bold the header
        ws.cell(row=1, column=3).font = Font(bold=True)
        # TODO clear out column 1 and 2 but have the user decide that with a boolean

        # go thru every row
        for x in range(2, n+1):
            # print(type(ws.cell(row=x, column=1).value), ws.cell(row=x, column=1).value)
            ws.cell(
                row=x, column=3).value = f"=HYPERLINK({get_column_letter(2)}{x},{get_column_letter(1)}{x})"
    wb.save(filename)


def processToExcelFile() -> str:
    # https://newsapi.org/docs/endpoints/top-headlines
    possibleCategories = ['business', 'entertainment',
                          'general', 'health', 'science', 'sports', 'technology']
    dfs = []
    if not os.path.isdir("./files"):
        os.mkdir('./files')
    filename = f'./files/news_{time.strftime("%Y%m%d-%H%M%S")}.xlsx'
    xlw = pd.ExcelWriter(filename)
    for category in possibleCategories:
        dfs.append(getNewsAsDataFrame(category))
    # print(os.listdir('./CSVs'))
    # merge them into one excel file separated into different sheets
    for i, df in enumerate(dfs):
        df.to_excel(xlw, sheet_name=possibleCategories[i], index=False)
    xlw.close()
    return filename


if __name__ == '__main__':
    pass
    filename = processToExcelFile()
    addHyperlinks(filename)



# region tkinter code
# ? window config
# root = tk.Tk()
# root.geometry('')
# root.title('News')
# scrollFrame = Frame(root)
# scrollFrame.pack(fill=BOTH, expand=1)
# canvas = Canvas(scrollFrame)
# canvas.pack(side=LEFT, fill=BOTH,expand=1)
# scrollbar = tk.Scrollbar(scrollFrame, orient=VERTICAL, command=canvas.yview)
# scrollbar.pack(side=RIGHT, fill=Y)
# #configure canvas for scrollbar
# canvas.configure(yscrollcommand=scrollbar.set)
# canvas.bind('<Configure>', lambda event: canvas.configure(scrollregion=canvas.bbox('all')))
# secondFrame = Frame(canvas,bg='gray')
# secondFrame.pack(padx=500,pady=500)
# canvas.create_window((0,0), window=secondFrame, anchor='nw')
# button = tk.Button(secondFrame, font=24, text='Reload', command=getNews,)
# button.pack()
# label = tk.Label(secondFrame, font=24, justify="left",anchor=NW, background='black', foreground='white')
# # pack on to the window
# label.pack(side=BOTTOM)
# label.config(text=getNews())
# thirdFrame = Frame(canvas,bg='gray')
# canvas.create_window((150,100), window=thirdFrame)
# root.mainloop()
# endregion